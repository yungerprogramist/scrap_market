import requests
from bs4 import BeautifulSoup
import lxml
import openpyxl


#api:      https://plati.io/api/search.ashx?query={поисковая_фраза}&pagesize={товаров на странице}&pagenum={номер страницы}&visibleOnly={только доступные}&response={формат ответа}
ALLOWED_GAMES = [ 
    'Albion', 
    'Apex',                           #todo перепроверить названия игр на сайте, иногда по полному названию находит не все, нужно поменять
    'ArcheAge', 
    'Ashes', 
    'Black Desert', 
    'Counter-Strike', 
    'Diablo II: Resurrected', 
    'Diablo Immortal', 
    'Diablo IV', 
    'Dota', 
    'Escape From Tarkov', 
    'Eve Online', 
    'Fortnite', 
    'Genshin', 
    'SAMP', 
    'League of Legends', 
    'Lineage II: Essence', 
    'Lineage II: Freeshards', 
    'Lineage II: Legacy', 
    'Lineage II: Main', 
    'Lost Ark', 
    'New World', 
    'Path of Exile', 
    'Perfect World', 
    'PUBG', 
    'PUBG Mobile', 
    'Raid Shadow Legends', 
    'Rust', 
    'Standoff 2', 
    'The Elder Scrolls Online', 
    'Throne and Liberty', 
    'Valorant', 
    'War Thunder', 
    'Warface', 
    'World of Tanks', 
    'World of Tanks Blitz', 
    'World of Warcraft Classic', 
    'World of Warcraft Dragonflight', 
    'World of Warcraft WotLK Classic', 
    'World of Warcraft Бесплатные серверы', 
    'Аллоды Онлайн', 
    ]


book= openpyxl.Workbook()
book.save('scrap.xlsx')

wb = openpyxl.load_workbook('scrap.xlsx')
ws = wb.active
ws.cell(row=1, column=1, value = 'продавец')
ws.cell(row=1, column=2, value = 'сделок') 
ws.cell(row=1, column=3, value = 'email')
ws.cell(row=1, column=4, value = 'telegram')
ws.cell(row=1, column=5, value = 'skype')
ws.cell(row=1, column=6, value = 'discord')
ws.cell(row=1, column=7, value = 'whatssap')
wb.save('scrap.xlsx')
#обозначени игр в голове таблице 
collumn_head_game = 8
for game in ALLOWED_GAMES:
    ws.cell (row = 1, column = collumn_head_game, value = game)
    collumn_head_game +=1 

# координатные переменные для таблицы
row_saller = 2 
column_saller = 1

row_solder = 2
column_solder = 2

row_email = 1
column_email = 3

row_tg = 2
column_tg = 4

row_skype = 2
column_skype = 5

row_discord = 2 
column_discord = 6

row_whatssap = 2
column_whatssap = 7

row_game = 2


column_user = 7

for game in ALLOWED_GAMES:
    response=requests.get(url=f'https://plati.io/api/search.ashx?query={game}&pagesize=500&visibleOnly=true&response=json')          #кидаем запрос

    print (game)

    json=response.json()       #переводим все в json формат
    data=json['items']              #получаем список игр
    seller_data={}                     #словарь типа ИМЯ ПРОДАВЦА: Количество лотов в нашей игре

    seller_contact = []
    
    for el in data: #бежим по ним циклом
        id=el['seller_id']                 #получаем id продавца
        seller_name=el['seller_name']   #получаем имя продавца
        if seller_name in seller_data:              #если имя продавца уже есть в списке лотов, то просто добавляем к количеству лотов 1
            seller_data[seller_name]+=1
        else:                                      #если нет, то добавляем его имя в словарь и парсим его страничку
            try:
                seller_data[seller_name]=1
                seller_response=requests.get(url=f'https://plati.market/seller/{seller_name}/{id}/')               #перехожу на страницу продавца
                soup=BeautifulSoup(seller_response.content,'lxml')                #варю суп
                sales=soup.find('div',class_='merchant-statistic').find('ol').find_all('li')                         #блок, чтобы достать количество продаж
                for sale in sales:
                    
                    if 'Sold' in sale.text:
                        qount_sales=sale.text.replace("Sold : ",'')
                        # print(f'юзер -- {seller_name}==== колво -- {qount_sales}')
                        break
                # print(qount_sales)                                                                                           #конец блока
                contacts=soup.find('div',class_='merchant-contacts').find_all('tr')                      #блок для поиска контактов
            except:
                continue
            
            # row_email +=1  #переносит строку емайла
            

            email_stop = False 
            teleg22 = False
            skype22 = False
            discord22 = False
            whatsapp22 = False
            for c_name in contacts:
                if 'E-mail:' in c_name.text:
                    
                    email=c_name.find('td').find('a').text
                    email_stop = True 
                    # print(email)
                    # ws.cell(row = row_email, column = column_email , value = email) #записывает емаил
                    # wb.save('scrap.xlsx')


                

            
                try:
                    img=c_name.find('th').find_all('img')
                    
                    

                    for c_image in img :
                        if c_image['src'] == '/images/telegram_128.png':
                            teleg = c_name.find('td').text #===========================================================

                            teleg22 = True 


                            # print(f'телеграм -- {teleg}')
                        # else:
                        #     teleg = 'None'
                        #     print(f'телеграмм -- {teleg}')
                            
                        if c_image['src'] == '/images/skype_128.png':
                            skype = c_name.find('td').text #============================================================

                            skype22 = True
                            # print(f'скайпS -- {skype}')
                        # else:
                        #     skype = 'None'
                        #     print(f'скайп -- {skype}')
                        
                        if c_image['src'] == '/images/discorduser.png':
                            discord = c_name.find('td').text #===============================================================

                            discord22 = True 
                            # print(f'дискорд -- {discord}')
                        # else:
                        #     discord = 'None'
                        #     print(f'дискорд -- {discord}')

                        if c_image['src'] == '/images/whatsapp_128.png':
                            whatsapp = c_name.find('td').text 

                            whatsapp22 = True 
                        # else:
                        #     whatsapp='None'                                           #конец блока
                    
                except:
                    pass

            stop = False
            
            
            for row_c in range (2, ws.max_row+1):
                polzov = ws.cell(row = row_c , column =1).value
                if polzov == seller_name:
                    stop = True

            # if [seller_name,email , whatsapp , teleg , skype , discord] not in seller_contact: 
            #     seller_contact.append([seller_name,email , whatsapp , teleg , skype , discord])


                    
            
            if not stop:
                ws.cell (row = row_saller, column = column_saller , value = seller_name)
                row_saller +=1 

                ws.cell(row = row_solder, column = column_solder , value = qount_sales)
                row_solder +=1 
                
                
                if teleg22 :
                    ws.cell(row = row_tg , column = column_tg, value = teleg)
                row_tg +=1 
                if skype22 :
                    ws.cell(row = row_skype, column =column_skype , value = skype )
                row_skype +=1 
                if discord22 :
                    ws.cell(row = row_discord, column = column_discord , value = discord )
                row_discord +=1 
                if whatsapp22:
                    ws.cell(row = row_whatssap, column = column_whatssap , value = whatsapp )
                row_whatssap +=1 

                row_email +=1

                if email_stop :
                    ws.cell(row = row_email, column = column_email , value = email)

                wb.save('scrap.xlsx')
    




            
    column_user +=1 

    keys = seller_data.keys()
    for row_count in range (2 , ws.max_row+1):
        user = ws.cell(row = row_count, column = 1).value
        for key in keys:
            try:
                if user in key:
                    
                    ws.cell(row = row_count , column = column_user, value = seller_data[key] )
                    wb.save('scrap.xlsx')
            except:
                pass 

