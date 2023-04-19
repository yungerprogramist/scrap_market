import requests 
from bs4 import BeautifulSoup 
import openpyxl

#не забудь раскоментить 45 строчку с добавлением продавцов в список 
def scrap():
    url = 'https://plati.market/games/'
    url_users = 'https://plati.market/seller'

    #список продавцов для обработки на неповторение 
    spisok_sellers = []
    spisok_sellers.clear

    #список парсящихся игр -----------------------------------------------------
    game_dict = [
    'Albion Online',
    'Apex Legends',
    'ArcheAge',
    'Ashes of Creation',
    'Black Desert Online',
    'Counter-Strike: Global Offensive',
    'Diablo II: Resurrected',
    'Diablo Immortal',
    'Diablo IV',
    'Dota 2',
    'Escape From Tarkov',
    'Eve Online',
    'Fortnite',
    'Genshin Impact',
    'GTA SAMP',
    'League of Legends',
    'Lineage II: Essence',
    'Lineage II: Freeshards',
    'Lineage II: Legacy',
    'Lineage II: Main',
    'Lost Ark',
    'New World',
    'Path of Exile',
    'Perfect World',
    'PUBG',
    'PUBG Mobile',
    'Raid: Shadow Legends',
    'Rust',
    'Standoff 2',
    'The Elder Scrolls Online',
    'Throne and Liberty',
    'Valorant',
    'War Thunder',
    'Warface',
    'World of Tanks',
    'World of Tanks: Blitz',
    'World of Warcraft: Classic',
    'World of Warcraft: Dragonflight',
    'World of Warcraft: WotLK Classic',
    'World of Warcraft: Бесплатные серверы',
    'Аллоды Онлайн',
    ]


    #создание обозначений в таблице и самой таблицы
    book= openpyxl.Workbook()
    book.save('scrap.xlsx')

    wb = openpyxl.load_workbook('scrap.xlsx')
    ws = wb.active
    ws.cell(row=1, column=1, value = 'продавец')
    ws.cell(row=1, column=2, value = 'сделок') 
    ws.cell(row=1, column=3, value = 'email')
    ws.cell(row=1, column=4, value = 'telegram')
    ws.cell(row=1, column=5, value = 'skype')
    ws.cell(row=1, column=6, value = 'discord')
    ws.cell(row=1, column=7, value = 'whatssap')
    wb.save('scrap.xlsx')
    #обозначени игр в голове таблице 
    collumn_head_game = 8
    for game in game_dict:
        ws.cell (row = 1, column = collumn_head_game, value = game)
        collumn_head_game +=1 

    # координатные переменные для таблицы
    row_saller = 2 
    column_saller = 1

    row_solder = 2
    column_solder = 2

    row_email = 1
    column_email = 3

    row_tg = 2
    column_tg = 4

    row_skype = 2
    column_skype = 5

    row_discord = 2 
    column_discord = 6

    row_whatssap = 2
    column_whatssap = 7

    row_game = 1

    response = requests.get(url).text 
    soup = BeautifulSoup(response , 'lxml')
    #div  в котором лежат все игры
    block_game = soup.find ('ul', class_ = 'titles-alphabet clearfix')

    for links_game in block_game:
        link_g_derty = links_game.find ('a').get('href')
        link_g = url + link_g_derty #готовые сылка на игру

        # print(link_g)

        name_game = links_game.find('a').text #названия игр 
        
    # for name_dict in  game_dict: #проверяем сходится ли название игры со списком игр для парсинга
        if name_game in game_dict:


            response = requests.get(link_g).text
            soup =BeautifulSoup (response, 'lxml') #варим суп уже этой странице с игрой

            block_card = soup.find('table' , class_ = 'goods-table goods-table-category')

            # for seller_block in block_card:

            seller_block1 = block_card.find_all('td', class_ = 'product-merchant') # достает все карточки таблицы товаров

            for info_card in seller_block1:
                link_seller_derty = info_card.find('a').get('href') 
                link_seller = url_users + link_seller_derty
                name_seller = info_card.find('a').text#имя продавца ================================================

                
                

                print('')
                print('')
                # print(f'{name_seller} -------------продавец') #-----------------------------------------------------------------
                # print(link_seller)


                if name_seller not in spisok_sellers: # проверка на повторение пользователей
                    spisok_sellers.append(name_seller) #добавляет пользователя в список для проверки на повторение в след цикле 

                    #сохранение в таблицу имен
                    ws.cell(row = row_saller, column = column_saller, value = name_seller)
                    wb.save('scrap.xlsx')
                    row_saller += 1

                    response = requests.get(link_seller).text
                    soup =BeautifulSoup (response, 'lxml') #варим суп уже этой странице с пользователем

                    user_block = soup.find ('div', class_='merchant-contacts')
                    
                    all_block = user_block.find_all('tr') #блок с контактными данными на одно приложение 

                    # print(all_block)
                    row_email +=1

                    for c_name in all_block: #в этом цикле достали все контактные данные----------------------------
                        app_block = c_name.find('th')
                        # print(app_block)

                        if app_block.text == 'E-mail:':
                            email = c_name.find('td').text #почта =============================================================
                            # print(f'почта -- {email}')

                            ws.cell(row = row_email, column = column_email, value = email)
                             
                            wb.save('scrap.xlsx')

                        else:
                            # row_email +=1
                            block_img = app_block.find_all('img') #это блок фотографии через него мы будем проверять приложение 
                            
                            for c_image in block_img: #
                                if c_image['src'] == '/images/telegram_128.png':
                                    teleg = c_name.find('td').text #===========================================================
                                    # print(f'телеграм -- {teleg}')
                                # else:
                                #     teleg = 'None'
                                #     print(f'телеграмм -- {teleg}')
                                    
                                if c_image['src'] == '/images/skype_128.png':
                                    skype = c_name.find('td').text #============================================================
                                    # print(f'скайпS -- {skype}')
                                # else:
                                #     skype = 'None'
                                #     print(f'скайп -- {skype}')
                                
                                if c_image['src'] == '/images/discorduser.png':
                                    discord = c_name.find('td').text #===============================================================
                                    # print(f'дискорд -- {discord}')
                                # else:
                                #     discord = 'None'
                                #     print(f'дискорд -- {discord}')

                                if c_image['src'] == '/images/whatsapp_128.png':
                                    whatsapp = c_name.find('td').text #===============================================================
                                    # print(f'ватсап -- {whatsapp}')
                                # else:
                                #     whatsapp = 'None'
                                #     print(f'ватсап -- {whatsapp}')

                        
                    #кол-во проданных
                    sell_block = soup.find('div', class_ = 'merchant-statistic')
                    solder1 = sell_block.find('ol')
                    solder2 = solder1.find('li').text
                    solder3 = solder2.split(' ')

                    solder = solder3[2] #================================================================================
                    #таблица 
                    ws.cell(row = row_solder , column = column_solder , value = solder)
                    row_solder +=1 
                    wb.save('scrap.xlsx')

                    # print (f'!!!сделок!!! -- {solder}')


                    # кол-во лотов в играх
                    item_block = soup.find('div', class_='sort_by')
                    item_block_all = item_block.find('optgroup', style = 'font-style:normal;font-weight:normal;')

                    row_game +=1 #перемещение на другую строку (следующего пользователя) 

                    for item in item_block_all:
                        # print(item)
                        try:
                            game_derty = item.text 
                            # print(game_derty)
                            recycle = game_derty.replace('Games >> ', '')
                            recycle1 = recycle.replace (')', '')
                            recycle2 = recycle1.split('(') # тут храниться два значения - название и количество
                            # print(recycle)
                            # print(recycle1)
                            # print(recycle2)

                            lot_game_name = recycle2[0]  #тут лот игр с названиями
                            lot_game_number = recycle2[1] #тут упорядоченные числа лотов 
                            # print (f'лотов в игре -- {lot_game1} {lot_game2}')
                            

                            # print (lot_game_name)

                            for col in range(7, ws.max_column+1):
                                game_from_dict = ws.cell(row = 1, column = col).value
                                # print (game_from_dict)
                                if (game_from_dict in lot_game_name) or (lot_game_name in game_from_dict)  :
                                    # print('прошло--------------------------------')
                                    ws.cell (row = row_game, column = col, value = lot_game_number)
                                    wb.save('scrap.xlsx')

# Albion Online
# Apex Legends
# ArcheAge
# Ashes of Creation
# Black Desert Online
# Counter-Strike: Global Offensive
# Diablo II: Resurrected
# Diablo Immortal
# Diablo IV
# Dota 2
# Escape From Tarkov
# Eve Online
# Fortnite
# Genshin Impact
# GTA SAMP
# League of Legends
                                


                                

                            # for x in range (1,ws.max_column+1):
                            #     values = ws.cell(row = 1, column = x).value
                            #     if values == 3:
                            #         ws.cell(row=2, column=x, value = 'gfh')
                            #         wb.save('123.xlsx')

                            '''
                            логика распределения игр:
                            1. сверяется есть ли такая игра в списке 
                            2. ищет ее позицию в КОЛОНКЕ 
                            3. подставляет цифру в найденную позицию колонки
                             
                            '''
                            

                        except:
                            continue
                        # Games >> Aion (3)

                else:
                    continue


def main():
    scrap()

if __name__ == "__main__":
    main()

